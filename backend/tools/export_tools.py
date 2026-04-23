"""Exportación profesional: PDF y Excel para contadores."""

import io
from datetime import date, datetime
from typing import Optional


def _get_month_transactions(year: int, month: int):
    from sqlmodel import Session, select
    from backend.db.database import engine
    from backend.models.transaction import Transaction
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    with Session(engine) as session:
        return session.exec(
            select(Transaction)
            .where(Transaction.date >= str(start))
            .where(Transaction.date < str(end))
            .order_by(Transaction.date)
        ).all()


def export_excel(year: int = 0, month: int = 0) -> bytes:
    """Generates an Excel workbook with transactions and summary. Returns bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from backend.models.transaction import TransactionType

    today = date.today()
    y = year or today.year
    m = month or today.month
    txs = _get_month_transactions(y, m)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transactions ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transacciones"

    header_fill = PatternFill("solid", fgColor="1C6B3A")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["ID", "Fecha", "Tipo", "Categoría", "Descripción", "Comercio", "Monto (COP)"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_income = 0.0
    total_expenses = 0.0

    for row_idx, tx in enumerate(txs, 2):
        ws.cell(row=row_idx, column=1, value=tx.id)
        ws.cell(row=row_idx, column=2, value=tx.date.strftime("%d/%m/%Y"))
        ws.cell(row=row_idx, column=3, value="Ingreso" if tx.type == TransactionType.income else "Gasto")
        ws.cell(row=row_idx, column=4, value=tx.category)
        ws.cell(row=row_idx, column=5, value=tx.description)
        ws.cell(row=row_idx, column=6, value=tx.merchant or "")
        amount_cell = ws.cell(row=row_idx, column=7, value=tx.amount)
        amount_cell.number_format = '#,##0'

        if tx.type == TransactionType.income:
            total_income += tx.amount
            ws.cell(row=row_idx, column=3).font = Font(color="16A34A")
        else:
            total_expenses += tx.amount
            ws.cell(row=row_idx, column=3).font = Font(color="DC2626")

    # Column widths
    widths = [6, 12, 10, 15, 40, 25, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    month_name = datetime(y, m, 1).strftime("%B %Y")
    ws2["A1"] = f"Resumen Financiero — {month_name}"
    ws2["A1"].font = Font(size=14, bold=True, color="1C6B3A")
    ws2.merge_cells("A1:B1")

    summary_rows = [
        ("Total Ingresos", total_income),
        ("Total Gastos", total_expenses),
        ("Balance", total_income - total_expenses),
        ("N° Transacciones", len(txs)),
    ]
    for r_idx, (label, value) in enumerate(summary_rows, 3):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        cell = ws2.cell(row=r_idx, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '"$"#,##0'
            if label == "Balance":
                cell.font = Font(color="16A34A" if value >= 0 else "DC2626")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_pdf(year: int = 0, month: int = 0, mode: str = "standard") -> bytes:
    """Generates a professional PDF report. Returns bytes.

    Args:
        mode: 'standard' for regular report, 'contador' for accountant-ready format.
    """
    from fpdf import FPDF
    from backend.models.transaction import TransactionType

    today = date.today()
    y = year or today.year
    m = month or today.month
    txs = _get_month_transactions(y, m)
    month_name = datetime(y, m, 1).strftime("%B %Y")

    total_income = sum(t.amount for t in txs if t.type == TransactionType.income)
    total_expenses = sum(t.amount for t in txs if t.type == TransactionType.expense)
    balance = total_income - total_expenses

    by_category: dict[str, float] = {}
    for t in txs:
        if t.type == TransactionType.expense:
            by_category[t.category] = by_category.get(t.category, 0) + t.amount

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Header
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(28, 107, 58)  # Leaf green
    pdf.cell(0, 12, "Leaf", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f"Reporte Financiero {'— Modo Contador' if mode == 'contador' else ''}", ln=True, align="C")
    pdf.cell(0, 6, month_name, ln=True, align="C")
    pdf.ln(6)

    # Summary box
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, "Resumen del Mes", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(22, 163, 74)
    pdf.cell(60, 7, "Ingresos:")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, f"${total_income:,.0f}", ln=True)
    pdf.set_text_color(220, 38, 38)
    pdf.cell(60, 7, "Gastos:")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, f"${total_expenses:,.0f}", ln=True)
    r, g, b = (22, 163, 74) if balance >= 0 else (220, 38, 38)
    pdf.set_text_color(r, g, b)
    pdf.cell(60, 7, "Balance:")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, f"${balance:+,.0f}", ln=True)
    pdf.ln(4)

    # Category breakdown
    if by_category:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, "Gastos por Categoria", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for cat, total in sorted(by_category.items(), key=lambda x: -x[1]):
            pct = total / total_expenses * 100 if total_expenses > 0 else 0
            pdf.cell(70, 6, cat.capitalize())
            pdf.cell(40, 6, f"${total:,.0f}")
            pdf.cell(0, 6, f"{pct:.1f}%", ln=True)
        pdf.ln(4)

    # Transactions table
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, "Detalle de Transacciones", ln=True)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(28, 107, 58)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(22, 7, "Fecha", fill=True)
    pdf.cell(18, 7, "Tipo", fill=True)
    pdf.cell(28, 7, "Categoria", fill=True)
    pdf.cell(80, 7, "Descripcion", fill=True)
    pdf.cell(0, 7, "Monto", fill=True, ln=True)

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(0, 0, 0)
    for i, tx in enumerate(txs):
        pdf.set_fill_color(240, 240, 240) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        fill = i % 2 == 0
        pdf.cell(22, 6, tx.date.strftime("%d/%m/%Y"), fill=fill)
        pdf.cell(18, 6, "Ingreso" if tx.type == TransactionType.income else "Gasto", fill=fill)
        pdf.cell(28, 6, tx.category[:14], fill=fill)
        pdf.cell(80, 6, tx.description[:40], fill=fill)
        pdf.cell(0, 6, f"${tx.amount:,.0f}", fill=fill, ln=True)

    if mode == "contador":
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(28, 107, 58)
        pdf.cell(0, 10, "Notas para el Contador", ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(0, 0, 0)
        notes = [
            f"• Período: {month_name}",
            f"• Total transacciones registradas: {len(txs)}",
            f"• GMF estimado (0.4% sobre gastos): ${total_expenses * 0.004:,.0f}",
            f"• GMF deducible (50%): ${total_expenses * 0.002:,.0f}",
            "• Los ingresos y gastos provienen del registro en Leaf.",
            "• Verificar con extractos bancarios para conciliación.",
            "• Para declaración de renta solicitar Formulario 210 en Leaf.",
        ]
        for note in notes:
            pdf.cell(0, 7, note, ln=True)

    return bytes(pdf.output())
